from __future__ import annotations

"""Excel (.xlsx) hisobot generatori — openpyxl.

Barcha hisobotlar bitta format: har varaq (sheet) = (nom, sarlavhalar, qatorlar).
Qiymatlar: str/int/float/Decimal/datetime — openpyxl o'zi yozadi."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4A6FA5", end_color="4A6FA5", fill_type="solid")
_MAX_SHEET_NAME = 31


def _cell_value(value: object) -> object:
    if value is None:
        return "-"
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def build_xlsx(sheets: list[tuple[str, list[str], list[list[object]]]]) -> bytes:
    """sheets: [(varaq nomi, [ustun sarlavhalari], [[qator qiymatlari], ...]), ...]"""
    wb = Workbook()
    wb.remove(wb.active)

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title[:_MAX_SHEET_NAME] or "Sheet")
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        widths = [len(str(h)) for h in headers]
        for row in rows:
            values = [_cell_value(v) for v in row]
            ws.append(values)
            for i, v in enumerate(values[: len(widths)]):
                widths[i] = max(widths[i], min(len(str(v)), 60))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w + 3

    if not wb.sheetnames:
        wb.create_sheet(title="Empty")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
